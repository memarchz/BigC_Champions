from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import cloudinary
import cloudinary.uploader
import openpyxl
from openpyxl import load_workbook
import os, json, re
from datetime import datetime
from typing import Optional
import httpx

app = FastAPI(title="BigC Champion API")

# CORS — อนุญาต GitHub Pages หรือทุก origin ระหว่าง dev
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ============================================================
# CONFIG — ใส่ค่าจริงใน Railway environment variables
# ============================================================
CLOUDINARY_CLOUD_NAME = os.getenv("CLOUDINARY_CLOUD_NAME", "")
CLOUDINARY_API_KEY    = os.getenv("CLOUDINARY_API_KEY", "")
CLOUDINARY_API_SECRET = os.getenv("CLOUDINARY_API_SECRET", "")

# OneDrive direct-download links ของ 2 ไฟล์ Excel
# วิธีได้ลิ้ง: เปิดไฟล์ใน OneDrive → Share → Copy link → เปลี่ยน ?e=xxx ท้าย URL เป็น download=1
RESPOND_EXCEL_URL  = "LOCAL_TEST"   # BigC_Champion_Respond.xlsx
DATABASE_EXCEL_URL = "LOCAL_TEST"  # ไฟล์ employee database

LOCAL_RESPOND_PATH  = "Test.xlsx"
LOCAL_DATABASE_PATH = "Test.xlsx"
# ============================================================

cloudinary.config(
    cloud_name=CLOUDINARY_CLOUD_NAME,
    api_key=CLOUDINARY_API_KEY,
    api_secret=CLOUDINARY_API_SECRET,
    secure=True,
)


# ---- helpers -----------------------------------------------

async def download_excel(url: str, dest: str):
    """ดาวน์โหลด Excel จาก OneDrive มาเก็บที่ /tmp"""
    async with httpx.AsyncClient(follow_redirects=True, timeout=30) as client:
        r = await client.get(url)
        r.raise_for_status()
        with open(dest, "wb") as f:
            f.write(r.content)


def excel_serial_to_date(serial) -> str:
    """แปลง Excel date serial เป็น YYYY-MM-DD"""
    if isinstance(serial, (int, float)):
        base = datetime(1899, 12, 30)
        delta = __import__("datetime").timedelta(days=int(serial))
        return (base + delta).strftime("%Y-%m-%d")
    return str(serial)


# ---- routes ------------------------------------------------

@app.get("/")
def root():
    return {"status": "BigC Champion API is running 🚀"}


@app.get("/employee/{emp_id}")
async def get_employee(emp_id: str):
    try:
        import openpyxl
        import os
        from fastapi import HTTPException

        excel_filename = "Test.xlsx"

        if not os.path.exists(excel_filename):
            print(f"❌ หาไฟล์ {excel_filename} ไม่เจอในโฟลเดอร์โครงการ!")
            raise HTTPException(status_code=404, detail="ไม่พบไฟล์ Excel รายชื่อพนักงาน")

        wb = openpyxl.load_workbook(excel_filename, data_only=True)
        
        # 🟢 พี่ตั้งชื่อแท็บเผื่อไว้ให้ ถ้าของหนูชื่อแท็บ "Employee" หรือชื่ออื่น สามารถเปลี่ยนตรงนี้ได้เลยนะครับ
        sheet_name = "Employee" 
        if sheet_name not in wb.sheetnames:
            sheet_name = wb.sheetnames[0] # ถ้าชื่อไม่ตรง ให้มันดึงแท็บแรกสุดของไฟล์มาใช้แทนเพื่อความปลอดภัย
            
        ws = wb[sheet_name]

        # เริ่มวิ่งไล่ตรวจทีละแถว (ข้ามหัวตารางแถวที่ 1)
        for row in ws.iter_rows(min_row=2, values_only=True):
            # ตรวจสอบว่าแถวนั้นมีข้อมูล และคอลัมน์ C (index 2) ตรงกับรหัสพนักงานที่พิมพ์มาไหม
            if len(row) >= 4 and row[2] is not None and str(row[2]).strip() == emp_id.strip():
                print(f"🟢 เจอข้อมูลแล้ว! พนักงาน: {row[3]} สังกัดสาขา: {row[1]}")
                return {
                    "employeeId": str(row[2]).strip(),   # คอลัมน์ C = รหัสพนักงาน
                    "employeeName": str(row[3]).strip(), # คอลัมน์ D = ชื่อพนักงาน
                    "storeCode": str(row[0]).strip(),    # คอลัมน์ A = รหัสสาขา
                    "storeName": str(row[1]).strip()     # คอลัมน์ B = ชื่อสาขา
                }

        print(f"🔍 พยายามค้นหาเลข {emp_id} ในคอลัมน์ C แล้วแต่ไม่พบในตาราง")
        raise HTTPException(status_code=404, detail="ไม่พบรหัสพนักงานนี้ในระบบ")

    except HTTPException as http_err:
        raise http_err
    except Exception as e:
        print(f"❌ เกิดข้อผิดพลาดฝั่งระบบ Excel: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/upload")
async def upload_photo(
    file: UploadFile = File(...),
    slot: str = Form(...),
    storeCode: str = Form(...),
):
    """อัปโหลดรูปไป Cloudinary แล้วคืน URL"""
    try:
        contents = await file.read()
        folder   = f"bigc_champion/{storeCode}"
        public_id = f"{storeCode}_{slot}_{int(datetime.now().timestamp())}"

        result = cloudinary.uploader.upload(
            contents,
            folder=folder,
            public_id=public_id,
            resource_type="image",
            overwrite=True,
        )
        return {"url": result["secure_url"], "slot": slot}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


class SubmitPayload(BaseModel):
    storeCode:    str
    storeName:    str
    employeeId:   str
    employeeName: str
    date:         str
    photos:       dict          # { "HCL1.1": "https://...", ... }
    freqQuestion: Optional[str] = ""
    hasBgc:       Optional[bool] = False


@app.post("/submit")
async def submit_form(payload: dict):
    try:
        import openpyxl
        import os
        from fastapi import HTTPException
        from datetime import datetime

        excel_path = "Test.xlsx"
        if not os.path.exists(excel_path):
            raise HTTPException(status_code=404, detail="ไม่พบไฟล์ Test.xlsx ในโฟลเดอร์")

        wb = openpyxl.load_workbook(excel_path)
        sheet_name = "Respond"
        if sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=404, detail=f"ไม่พบแท็บที่ชื่อว่า {sheet_name} ในไฟล์ Excel")
        ws = wb[sheet_name]

        # 1. ดึงข้อมูลพื้นฐานจากหน้าเว็บ
        store_code = payload.get("storeCode", "")
        store_name = payload.get("storeName", "")
        emp_id = payload.get("employeeId", "")
        emp_name = payload.get("employeeName", "")
        visit_date = payload.get("date", "")  # รูปแบบ "YYYY-MM-DD"
        freq_q = payload.get("freqQuestion", "")
        has_bgc = "YES" if payload.get("hasBgc") else "NO"

        # ดึงปีและเดือนจากวันที่ตรวจ
        year_val = ""
        month_val = ""
        if visit_date:
            try:
                date_obj = datetime.strptime(visit_date, "%Y-%m-%d")
                year_val = date_obj.year    # เช่น 2026
                month_val = date_obj.month  # เช่น 6
            except:
                pass

        # ดึงลิงก์รูปภาพทั้งหมดที่ถูกส่งมาจากหน้าเว็บ
        photos = payload.get("photos", {})

        # 🟢 เริ่มจัดแถวข้อมูล (new_row) ไล่ทีละคอลัมน์ตามไฟล์ Excel จริงของหนู
        
        # คอลัมน์ A - E (ลำดับ 1-5)
        new_row = [
            store_code,    # A
            store_name,    # B
            emp_id,        # C
            emp_name,      # D
            visit_date     # E
        ]

        # คอลัมน์ F - N (ลำดับ 6-14): รูป HCL แบบมีจุด (HCL1.1 ถึง HCL3.3)
        hcl_dots = [
            "HCL1.1", "HCL1.2", "HCL1.3",
            "HCL2.1", "HCL2.2", "HCL2.3",
            "HCL3.1", "HCL3.2", "HCL3.3"
        ]
        for slot in hcl_dots:
            new_row.append(photos.get(slot, ""))

        # คอลัมน์ O - AC (ลำดับ 15-29): รูป BGC แบบมีจุด (BGC1.1 ถึง BGC5.3)
        bgc_dots = [
            "BGC1.1", "BGC1.2", "BGC1.3",
            "BGC2.1", "BGC2.2", "BGC2.3",
            "BGC3.1", "BGC3.2", "BGC3.3",
            "BGC4.1", "BGC4.2", "BGC4.3",
            "BGC5.1", "BGC5.2", "BGC5.3"
        ]
        for slot in bgc_dots:
            new_row.append(photos.get(slot, ""))

        # คอลัมน์ AD - AE (ลำดับ 30-31)
        new_row.append(freq_q)   # AD (Freq.Qs.)
        new_row.append(has_bgc)  # AE (HAS_BGC)

        # คอลัมน์ AF - AN (ลำดับ 32-40): รูป HCL แบบไม่มีจุด (HCL11 ถึง HCL33)
        # เผื่อหน้าเว็บส่งมาทั้งสองแบบ หรือกันเหนียว พี่จะเช็กทั้งแบบมีจุดและไม่มีจุดให้เลยครับ
        hcl_no_dots = [
            ("HCL11", "HCL1.1"), ("HCL12", "HCL1.2"), ("HCL13", "HCL1.3"),
            ("HCL21", "HCL2.1"), ("HCL22", "HCL2.2"), ("HCL23", "HCL2.3"),
            ("HCL31", "HCL3.1"), ("HCL32", "HCL3.2"), ("HCL33", "HCL3.3")
        ]
        for slot_no, slot_dot in hcl_no_dots:
            # ถ้ามีค่าส่งมาจากชื่อไหนก็นำค่านั้นมาใส่
            val = photos.get(slot_no, photos.get(slot_dot, ""))
            new_row.append(val)

        # คอลัมน์ AO - BC (ลำดับ 41-55): รูป BGC แบบไม่มีจุด (BGC11 ถึง BGC53)
        bgc_no_dots = [
            ("BGC11", "BGC1.1"), ("BGC12", "BGC1.2"), ("BGC13", "BGC1.3"),
            ("BGC21", "BGC2.1"), ("BGC22", "BGC2.2"), ("BGC23", "BGC2.3"),
            ("BGC31", "BGC3.1"), ("BGC32", "BGC3.2"), ("BGC33", "BGC3.3"),
            ("BGC41", "BGC4.1"), ("BGC42", "BGC4.2"), ("BGC43", "BGC4.3"),
            ("BGC51", "BGC5.1"), ("BGC52", "BGC5.2"), ("BGC53", "BGC5.3")
        ]
        for slot_no, slot_dot in bgc_no_dots:
            val = photos.get(slot_no, photos.get(slot_dot, ""))
            new_row.append(val)

        # คอลัมน์ BD - BE (ลำดับ 56-57): สองช่องสุดท้ายของตาราง
        new_row.append(year_val)   # BD (Year)
        new_row.append(month_val)  # BE (Month)

        # 3. บันทึกข้อมูลและกดเซฟไฟล์ลงเครื่องคอมพิวเตอร์
        ws.append(new_row)
        wb.save(excel_path)
        
        print(f"🟢 จัดคอลลัมน์เป๊ะแล้ว! บันทึกข้อมูลของพนักงานรหัส {emp_id} ลงช่องถูกต้องเรียบร้อย")
        return {"status": "success", "message": "บันทึกข้อมูลลงตาราง Excel ล็อกถูกต้องเรียบร้อยแล้วค่ะ"}
        
    except Exception as e:
        print(f"❌ เออเรอร์ตอนเซฟข้อมูล: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
